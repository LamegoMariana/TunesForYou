# Funciones Tunes For You

import webbrowser
import time

# Se hacen variables globales de listas con las opciones a elegir por el usuario para ahorrar líneas de código
val1 = ["A", "B", "C", "D"]
val = ["A", "B", "C"]
valink = ["A", "B"]

# Se hacen variables globales de listas vacías para insertar los subgéneros
listametal = []
listapop = []
listarock = []
listaelectro = []
listahip = []

# Se hacen variables globales de listas vacías para insertar los resultados obtenidos
listaresmetal = []
listarespop = []
listaresrock = []
listareselectro = []
listareship = []


# Bienvenida y pregunta para Inicio de sesión
def saludo():
    respuesta = ""
    # Se da formato al título
    header = "┌─ ♪ ────────────────────────────────┐"
    print("\n", header.center(48), "\n")
    tittle = "¸♬¸ Bienvenidos a Tunes For You ¸♬¸"
    print(tittle.center(49), "\n")
    header2 = "└──────────────────────────────── ♪ ─┘"
    print(header2.center(50), "\n")
    # Se pregunta si el usuario ya tiene cuenta (datos registrados) y se valida la respuesta
    while respuesta not in valink:
        print("¿Ya tienes una cuenta? A = Si | B = No", end="")
        respuesta = input(": ")
        if respuesta not in valink:
            print("Respuesta desconocida, sólo introduce 'A' o 'B'")
    return respuesta


# Leer nombres de usuarios del archivo
def leernombreusuario():
    from openpyxl import load_workbook
    filesheet = "./BaseDeDatos.xlsx"
    wb = load_workbook(filesheet)
    # Se lee la hoja 2, "Usuarios" (índice 1)
    wb.active = 1
    sheet = wb.active
    # Lee columna B (nombres de usuario) de la hoja 2 y hace una lista con esos valores
    col = sheet["B"]
    usuarios_lista = []
    for c in col:
        usuarios_lista.append(c.value)
    wb.close()
    return usuarios_lista


# Leer contraseñas de usuarios del archivo
def leercontrausuario():
    from openpyxl import load_workbook
    filesheet = "./BaseDeDatos.xlsx"
    wb = load_workbook(filesheet)
    # Se lee la hoja 2, "Usuarios" (índice 1)
    wb.active = 1
    sheet = wb.active
    # Lee columna C (contraseñas) de la hoja 2 y hace una lista con esos valores
    colu = sheet["C"]
    contras_lista = []
    for c in colu:
        contras_lista.append(c.value)
    wb.close()
    return contras_lista


# Inicio de sesión cuando ya se tiene cuenta (leer datos, usuario y contraseña, del archivo)
def inicio_sesion():
    # Se da formato al título
    tittle = "¸♬¸ INICIO DE SESIÓN ¸♬¸"
    print("\n ¡Genial!")
    print(tittle.center(50))
    # Se pide el nombre de usuario
    print("\nIngresa tu nombre de usuario o escribe 'X' para salir", end="")
    nomu = input(": ")
    # Con "X" se termina el programa
    if nomu == "X":
        print("\n¸♬¸ Gracias por usar Tunes For You ¸♬¸")
        print("¡Vuelve pronto!")
        exit()
    else:
        # Ciclo para volver a pedir nombre de usuario en caso de ser demasiado corto
        while len(nomu) <= 1:
            print("Los nombres de usuario son más largos.Inténtalo de nuevo.")
            nomu = input("\tIngresa tu nombre de usuario: ")
        # Se lee la lista con los nombres de usuario y se compara para ver si existe el que se ingresó
        leernombreusuario()
        # Si no existe se le hace saber al usuario
        while nomu not in leernombreusuario():
            print("\nNombre de usuario inexistente.")
            # El usuario  puede registrar el nombre con la función "Crear cuenta"
            nomu = input("\t¿Quieres registrarlo? 'A' = Si | 'B' = No: ")
            if nomu == "A":
                crearcuenta()
                break
            else:
                # Si no la registra, debe volver a intentar escribir su nombre o salir del programa
                print("Debes tener una cuenta. Inténtalo de nuevo o escribe 'X' para salir.")
                nomu = input("\tIngresa tu nombre de usuario: ")
                # Con "X" se termina el programa
                if nomu == "X":
                    print("\n¸♬¸ Gracias por usar Tunes For You ¸♬¸")
                    print("¡Vuelve pronto!")
                    exit()
                # Ciclo para volver a pedir nombre de usuario en caso de ser demasiado corto
                while len(nomu) <= 1:
                    print("Los nombres de usuario son más largos.Inténtalo de nuevo.")
                    nomu = input("\tIngresa tu nombre de usuario: ")
        # Si el nombre existe
        if nomu in leernombreusuario():
            print("\nNombre confirmado :D")
            # Se busca el índice en que se encuentra dentro de la lista para poder leer esa contraseña
            renglonusuario = leernombreusuario().index(nomu)
            # Se pide la contraseña
            print("\tIngresa tu contraseña", end="")
            contra = eval(input(": "))
            leercontrausuario()
            # Se comprueba que la contraseña ingresada coincida con la que corresponde al nombre de usuario ingresado
            contrausu = leercontrausuario()[renglonusuario]
            while contra != contrausu:
                print("Contraseña no válida para este usuario. Inténtalo de nuevo D:\n")
                contra = eval(input("Ingresa tu contraseña: "))
            if contra == contrausu:
                print("\nContraseña confirmada :D")
            # Se va al menú principal
                menu()


# Crear cuenta (grabar nuevos datos, de usuario y contraseña, en el archivo)
def crearcuenta():
    # Se da formato al título
    tittle = "\n¸♬¸ CREAR CUENTA ¸♬¸"
    print(tittle.center(50), "\n")
    # Se pide y valida un nombre de usuario mayor a un carácter
    print("Crea tu nombre de usuario o 'X' para salir.")
    usuario = input("Debe contener más de un caracter: ")
    if usuario == "X":
        print("\nGracias por usar Tunes For You. ¡Vuelve pronto!")
        exit()
    while not len(usuario) > 1:
        print("Nombre de usuario demasiado corto.Inténtalo de nuevo.")
        usuario = input("\tIngresa un nombre de usuario: ")
    # Se pide y valida una contraseña mayor a dos caracteres
    print("\nCrea tu contraseña.")
    contrasena = int(input("Debe contener más de dos dígitos y ser numérica: "))
    while not len(str(contrasena)) > 2:
        print("Contraseña demasiado corta.Inténtalo de nuevo.")
        contrasena = int(input("\tIngresa una contraseña numérica: "))
    # Si se cumplen las dos condiciones, se continúa el proceso
    if len(usuario) > 1 and len(str(contrasena)) > 2:
        # Se abre la base de datos
        from openpyxl import load_workbook
        filesheet = "./BaseDeDatos.xlsx"
        wb = load_workbook(filesheet)
        # Se lee la hoja 2, "Usuarios" (índice 1)
        wb.active = 1
        sheet = wb.active
        # Se busca cuál el último renglón para escribir en él
        siguienteinsert = sheet.max_row
        # Se crea una lista para ingresar los datos
        usuariostotal = []
        usuariostotal.append(siguienteinsert)
        usuariostotal.append(usuario)
        usuariostotal.append(contrasena)
        # Se graban los datos en el archivo
        try:
            sheet.append(usuariostotal)
            wb.save("./BaseDeDatos.xlsx")
            print("\nCuenta registrada exitosamente :D")
        except:
            print("Ocurrió un error al escribir :/")
        finally:
            wb.close()
            # Se va al menú principal
            menu()


# Menú Principal
def menu():
    # Se le da formato al título
    header = "┌─ ♪ ────────────────────────────────┐"
    print("\n", header.center(48), "\n")
    titulo = "¸♬¸ Menú principal ¸♬¸"
    print(titulo.center(48), "\n")
    header2 = "└──────────────────────────────── ♪ ─┘"
    print(header2.center(50), "\n")
    # Se despliegan las opciones disponibles
    print("\tA: Iniciar sistema de recomendación")
    print("\tB: Ver recomendaciones anteriores")
    print("\tC: Agregar música a la base de datos")
    print("\tD: Salir\n")
    Respuesta = input("¿Qué quieres hacer?: ")
    # Si la opción ingresada no coincide con las disponibles se vuelve a pedir
    while not ("A" <= Respuesta <= "D"):
        print("Respuesta desconocida, introduce sólo A, B, C ó D\n")
        Respuesta = input("¿Qué quieres hacer?: ")
    # Con A se inicia la función de 'Preguntas'
    if Respuesta == "A":
        print("\nIniciando Sistema de Recomendación")
        time.sleep(1)
        print(".", end="")
        time.sleep(1)
        print(".", end="")
        time.sleep(1)
        print(".", end="")
        time.sleep(1)
        preguntas()
    # Con B se inicia la función de 'Recomendaciones anteriores'
    if Respuesta == "B":
        recoant()
    # Con C se agrega música nueva, hay dos maneras
    if Respuesta == "C":
        print("\n¿Quieres agregar un género nuevo (A) o añadir a uno ya existente (B)?", end="")
        RespuestaC = input(": ")
        # Se valida que solo se ingrese 'A' o 'B'
        while not ("A" <= RespuestaC <= "B"):
            RespuestaC = input("Respuesta desconocida, sólo introduce 'A' o 'B': ")
        # Agregar datos completamente nuevos
        if RespuestaC == "A":
            ingresarmusicnueva1()
        # Agregar datos basados en existentes
        if RespuestaC == "B":
            ingresarmusicnueva2()
    # Con D se termina el programa
    if Respuesta == "D":
        print("\n¸♬¸ Gracias por usar Tunes For You ¸♬¸")
        print("¡Vuelve pronto!")
        exit()


# Inicia sistema de recomendación
def preguntas():
    # Se le da formato al título
    header = "┌─ ♪ ────────────────────────────────┐"
    print("\n", header.center(50), "\n")
    titulo = "¸♬¸ Tunes For You ¸♬¸"
    print(titulo.center(50), "\n")
    header2 = "└──────────────────────────────── ♪ ─┘"
    print(header2.center(50), "\n")
    print("Para cada pregunta, elige la opción que más se acerque a tu respuesta :D\n")
    print()
    # Se procede a Preguntas Metal
    preguntasMetal()


def preguntasMetal():
    print("")
    print("¿Te gusta escuchar canciones con guitarras eléctricas fuertes y distorsionadas,"
          " con ritmos enfáticos, enérgicos y potentes? \n")
    print("\tA) ¡Mucho!")
    print("\tB) Sí, me gusta")
    print("\tC) No mucho pero no me desagradan")
    print("\tD) Prefiero mejor otras canciones \n")
    MRespuesta1 = input()
    while MRespuesta1 not in val1:
        MRespuesta1 = input("Respuesta desconocida, introduce sólo A, B, C ó D\n")
    if MRespuesta1 == "A" or MRespuesta1 == "B":
        print("¿Cómo te sientes respecto a las canciones con ritmos calmados que evolucionan"
              " a melodías pesadas y viceversa?\n")
        print("\tA = ¡Las amo!")
        print("\tB = Me podrían gustar")
        print("\tC = Preferiría escuchar otra cosa\n")
        MRespuesta11 = input()
        while MRespuesta11 not in val:
            MRespuesta11 = input("Respuesta desconocida, introduce sólo A, B ó C\n")
        # Se agrega el subgénero a la lista del género
        if MRespuesta11 == "A":
            subgenero1 = "Progressive metal"
            listametal.append(subgenero1)
        # Se agrega el subgénero a la lista del género
        elif MRespuesta11 == "B":
            subgenero15 = "Progressive metal"
            listametal.append(subgenero15)

        elif MRespuesta11 == "C":
            print()
        print("¿Te podría interesar la fusión del metal y el folk tradicional, relacionado con el metal vikingo?\n")
        print("\tA = ¡Por supuesto!")
        print("\tB = Claro, ¿por qué no?")
        print("\tC = Mmm, quizás no\n")
        MRespuesta12 = input()
        while MRespuesta12 not in val:
            MRespuesta12 = input("Respuesta desconocida, introduce sólo A, B ó C\n")
        # Se agrega el subgénero a la lista del género
        if MRespuesta12 == "A":
            subgenero2 = "Folk metal"
            listametal.append(subgenero2)
        # Se agrega el subgénero a la lista del género
        elif MRespuesta12 == "B":
            subgenero25 = "Folk metal"
            listametal.append(subgenero25)
        elif MRespuesta12 == "C":
            print()
        print("¿En qué grado te gusta la música que combina el metal con otros géneros,"
              " como el hip hop, el grunge, el rock alternativo y el funk?\n")
        print("\tA = ¡Mucho!")
        print("\tB = Tal vez, suena interesante")
        print("\tC = No tanto\n")
        MRespuesta13 = input()
        while MRespuesta13 not in val:
            MRespuesta13 = input("Respuesta desconocida, introduce sólo A, B ó C\n")
        # Se agrega el subgénero a la lista del género
        if MRespuesta13 == "A":
            subgenero3 = "Nu metal"
            listametal.append(subgenero3)
        # Se agrega el subgénero a la lista del género
        elif MRespuesta13 == "B":
            subgenero35 = "Nu metal"
            listametal.append(subgenero35)

        elif MRespuesta13 == "C":
            print()
# Se envían a los usuarios a las recomendaciones de canciones y artistas,
        # en el caso que hayan seleccionado A o B en una de las preguntas
        if len(listametal) == 1:
            print("¡Muy bien! Te gustó el siguiente subgénero del metal:")
            print('\t, '.join(listametal))
            print("¡Te recomendaremos el siguiente artista con una canción! Sabemos que te encantará")
            print("")
            resultadosMetal()
        elif len(listametal) >= 2:
            print("¡Muy bien! Te recomendamos los siguientes subgéneros del metal:")
            print("", end="\t")
            print(', '.join(listametal))
            print("¡Te recomendamos los siguiente artistas con su determinado subgénero y"
                  " una de sus canciones! Sabemos que te encantarán :D")
            print("")
            resultadosMetal()

    # Si no hay valores en la lista o se seleccionó "C" o "D" en la preguntas inicial,
        # se pasa a las preguntas del siguiente género
        elif len(listametal) == 0:
            preguntasPop()

    elif MRespuesta1 == "C" or MRespuesta1 == "D":
        preguntasPop()


# Función con las recomendaciones resultantes de Metal
def resultadosMetal():
    # Se lee la base de datos en la primera hoja
    from openpyxl import load_workbook
    filesheet = "./BaseDeDatos.xlsx"
    wb = load_workbook(filesheet)
    wb.active = 0
    sheet = wb.active
    # Se imprime el título con estilo
    header = "┌─ ♪ ────────────────────────────────┐"
    print("\n", header.center(48), "\n")
    tittle = "¸♬¸ Resultados ¸♬¸"
    print(tittle.center(49), "\n")
    header2 = "└──────────────────────────────── ♪ ─┘"
    print(header2.center(50), "\n")
# Dependiendo al subgénero, se imprime el nombre del artista/grupo y una canción que se encuentre en la base de datos
    if "Progressive metal" in listametal:

        for metal1 in sheet.iter_rows(min_row=2, max_row=2, min_col=5, max_col=7, values_only=True):
            listaresmetal.append(metal1)
            res = " - ".join(metal1)
            print("\t- Progressive Metal: ", res)
    if "Folk metal" in listametal:
        for metal2 in sheet.iter_rows(min_row=5, max_row=5, min_col=5, max_col=7, values_only=True):
            listaresmetal.append(metal2)
            res = " - ".join(metal2)
            print("\t- Folk Metal: ", res)
    if "Nu metal" in listametal:
        for metal3 in sheet.iter_rows(min_row=8, max_row=8, min_col=5, max_col=7, values_only=True):
            listaresmetal.append(metal3)
            res = " - ".join(metal3)
            print("\t- Nu Metal: ", res)
    print("")
    SiguienteInsert = sheet.max_row
    print("¿Quieres que te llevemos directo al link de youtube de alguna canción?")
    print("\t A) ¡Sí!")
    print("\t B) No")

    # Se pregunta si se abre automáticamente el link de una canción a youtube
    reslink = input()
    while reslink not in val:
        reslink = input("Respuesta desconocida. Sólo introduce A ó B\n")
    if reslink == "A":
        if len(listametal) == 3:
            print("¡Perfecto :D! ¿De cuál subgénero quieres el link?")
            if "Progressive metal" and "Folk metal" and "Nu metal" in listametal:
                print("\tA) Progressive metal")
                print("\tB) Folk metal")
                print("\tC) Nu metal\t")
                reslink1 = input()
                while reslink1 not in val:
                    reslink1 = input("Respuesta desconocida. Sólo introduce A ó B\n")

                    # Se abre el link del subgénero seleccionado
                if reslink1 == "A":
                    prometal = listaresmetal[0]
                    webbrowser.open(prometal[2])
                    print("¡Esperemos que te guste!:D")
                elif reslink1 == "B":
                    prometal = listaresmetal[1]
                    webbrowser.open(prometal[2])
                    print("¡Esperemos que te guste!:D")
                elif reslink1 == "C":
                    prometal = listaresmetal[2]
                    webbrowser.open(prometal[2])
                    print("¡Esperemos que te guste!:D")
                print("¿Quieres seguir con el sistema de recomendación o volver al menú?")
                print("\t A) Seguir con el sistema")
                print("\t B) Menú principal")
                metalfin = input()
                while metalfin not in valink:
                    metalfin = input("Respuesta desconocida. Sólo introduce A ó B\n")
                if metalfin == "A":
                    preguntasPop()
                if metalfin == "B":
                    menu()
        elif len(listametal) == 2:
            print("¡Perfecto :D! ¿De cuál subgénero quieres el link?")
            if "Progressive metal" and "Folk metal" in listametal:
                print("\tA) Progressive metal")
                print("\tB) Folk metal")
                reslink1 = input()
                while reslink1 not in valink:
                    reslink1 = input("Respuesta desconocida. Sólo introduce A ó B\n")
                # Se abre el link del subgénero seleccionado
                if reslink1 == "A":
                    prometal = listaresmetal[0]
                    webbrowser.open(prometal[2])
                    print("¡Esperemos que te guste!:D")
                elif reslink1 == "B":
                    prometal = listaresmetal[1]
                    webbrowser.open(prometal[2])
                    print("¡Esperemos que te guste!:D")
                print("¿Quieres seguir con el sistema de recomendación o volver al menú?")
                print("\t A) Seguir con el sistema")
                print("\t B) Menú principal")
                metalfin = input()
                while metalfin not in valink:
                    metalfin = input("Respuesta desconocida. Sólo introduce A ó B\n")
                if metalfin == "A":
                    preguntasPop()
                if metalfin == "B":
                    menu()
            elif "Progressive metal" and "Nu metal" in listametal:
                print("\tA) Progressive metal")
                print("\tB) Nu metal")
                reslink1 = input()
                while reslink1 not in valink:
                    reslink1 = input("Respuesta desconocida. Sólo introduce A ó B\n")
                # Se abre el link del subgénero seleccionado
                if reslink1 == "A":
                    prometal = listaresmetal[0]
                    webbrowser.open(prometal[2])
                    print("¡Esperemos que te guste!:D")
                elif reslink1 == "B":
                    prometal = listaresmetal[1]
                    webbrowser.open(prometal[2])
                    print("¡Esperemos que te guste!:D")
                print("¿Quieres seguir con el sistema de recomendación o volver al menú?")
                print("\t A) Seguir con el sistema")
                print("\t B) Menú principal")
                metalfin = input()
                while metalfin not in valink:
                    metalfin = input("Respuesta desconocida. Sólo introduce A ó B\n")
                if metalfin == "A":
                    preguntasPop()
                if metalfin == "B":
                    menu()
            elif "Folk metal" and "Nu metal" in listametal:
                print("\tA) Folk metal")
                print("\tB) Nu metal")
                reslink1 = input()
                while reslink1 not in valink:
                    reslink1 = input("Respuesta desconocida. Sólo introduce A ó B\n")
                # Se abre el link del subgénero seleccionado
                if reslink1 == "A":
                    prometal = listaresmetal[0]
                    webbrowser.open(prometal[2])
                    print("¡Esperemos que te guste!:D")
                elif reslink1 == "B":
                    prometal = listaresmetal[1]
                    webbrowser.open(prometal[2])
                    print("¡Esperemos que te guste!:D")
                print("¿Quieres seguir con el sistema de recomendación o volver al menú?")
                print("\t A) Seguir con el sistema")
                print("\t B) Menú principal")
                metalfin = input()
                while metalfin not in valink:
                    metalfin = input("Respuesta desconocida. Sólo introduce A ó B\n")
                if metalfin == "A":
                    preguntasPop()
                if metalfin == "B":
                    menu()
        elif len(listametal) == 1:
            print("¡Perfecto :D! En un momento se abrirá")
            prometal = listaresmetal[0]
            webbrowser.open(prometal[2])
            print("¡Esperemos que te guste!:D")

    if reslink == "B":
        print()
    print("¿Quieres seguir con el sistema de recomendación o volver al menú?")
    print("\t A) Seguir con el sistema")
    print("\t B) Menú principal")
    metalfin = input()
    while metalfin not in valink:
        metalfin = input("Respuesta desconocida. Sólo introduce A ó B\n")
    # Dependiendo a la respuesta del usuario, se regresa al menú o se sigue con el sistema
    if metalfin == "A":
        preguntasPop()
    if metalfin == "B":
        menu()


# Función con las preguntas de Pop
def preguntasPop():
    print("")
    print("¿Dirías que la música que toma prestados elementos de otros estilos es de tu agrado?"
          "\n Considera que son canciones de corta a media duración y escritas en un formato básico, "
          "con uso habitual de estribillos repetidos. \n")
    print("A) ¡Mucho!")
    print("B) ¡Sí! Me gusta")
    print("C) No mucho pero no me desagrada")
    print("D) Prefiero mejor otro tipo de música \n")
    PRespuesta1 = input()
    while PRespuesta1 not in val1:
        PRespuesta1 = input("Respuesta desconocida, introduce sólo A, B, C ó D\n")
    if PRespuesta1 == "A" or PRespuesta1 == "B":
        print("¿Qué tanto te agradan las canciones con estilo suave, ensoñador y "
              "psicodélico que tienen la intención de sumergir al oyente en la melodía?\n")
        print("A = ¡Me encanta escucharlas!")
        print("B = Me agradan algo")
        print("C = Prefiero otra cosa\n")
        PRespuesta11 = input()
        while PRespuesta11 not in val:
            PRespuesta11 = input("Respuesta desconocida, introduce sólo A, B ó C\n")
        # Se agrega el subgénero a la lista del género
        if PRespuesta11 == "A":
            subgenero4 = "Dream pop"
            listapop.append(subgenero4)
        # Se agrega el subgénero a la lista del género
        elif PRespuesta11 == "B":
            subgenero45 = "Dream pop"
            listapop.append(subgenero45)

        elif PRespuesta11 == "C":
            print()
        print("¿Cómo te sientes respecto a la música que incorpora melodías vocales, "
              "pasajes energéticos, y melodía alegre?\n")
        print("A = ¡Uff, me encanta!")
        print("B = Podría escucharla de vez en cuando")
        print("C = Mmm, no me gustan\n")
        PRespuesta12 = input()
        while PRespuesta12 not in val:
            PRespuesta12 = input("Respuesta desconocida, introduce sólo A, B ó C\n")

        # Se agrega el subgénero a la lista del género
        if PRespuesta12 == "A":
            subgenero5 = "Power pop"
            listapop.append(subgenero5)

        # Se agrega el subgénero a la lista del género
        elif PRespuesta12 == "B":
            subgenero55 = "Power pop"
            listapop.append(subgenero55)

        elif PRespuesta12 == "C":
            print()
        print("¿Podría gustarte la música en otro idioma, que tiene influencias del "
              "rock de los sesentas, del jazz y de la electrónica?\n")
        print("A = Hai, mochiron! (¡Sí, por supuesto!)")
        print("B = Las escucho de vez en cuando")
        print("C = Prefiero algo que entienda\n")
        PRespuesta13 = input()
        while PRespuesta13 not in val:
            PRespuesta13 = input("Respuesta desconocida, introduce sólo A, B ó C\n")

        # Se agrega el subgénero a la lista del género
        if PRespuesta13 == "A":
            subgenero6 = "J-pop"
            listapop.append(subgenero6)

        # Se agrega el subgénero a la lista del género
        elif PRespuesta13 == "B":
            subgenero65 = "J-Pop"
            listapop.append(subgenero65)

        elif PRespuesta13 == "C":
            print()
        # Se envían a los usuarios a las recomendaciones de canciones y artistas,
        # en el caso que hayan seleccionado A o B en una de las preguntas
        if len(listapop) == 1:
            print("¡Muy bien! Te gustó el siguiente subgénero del pop:")
            print('\t, '.join(listapop))
            print("¡Te recomendaremos el siguiente artista con una canción! Sabemos que te encantará")
            print("")
            resultadosPop()
        elif len(listapop) >= 2:
            print("¡Muy bien! Te recomendamos los siguientes subgéneros del pop:")
            print("", end="\t")
            print(', '.join(listapop))
            print("¡Te recomendamos los siguiente artistas con su determinado subgénero y "
                  "una de sus canciones! Sabemos que te encantarán :D")
            print("")
            resultadosPop()
    # Si no hay valores en la lista o se seleccionó "C" o "D" en la preguntas inicial,
        # se pasa a las preguntas del siguiente género
        elif len(listapop) == 0:
            preguntasRock()
    elif PRespuesta1 == "C" or PRespuesta1 == "D":
        preguntasRock()


# Función con las recomendaciones resultantes de Pop
def resultadosPop():
    # Se lee la base de datos en la primera hoja
    from openpyxl import load_workbook
    filesheet = "./BaseDeDatos.xlsx"
    wb = load_workbook(filesheet)
    wb.active = 0
    sheet = wb.active
    # Se imprime el título con estilo
    header = "┌─ ♪ ────────────────────────────────┐"
    print("\n", header.center(48), "\n")
    tittle = "¸♬¸ Resultados ¸♬¸"
    print(tittle.center(49), "\n")
    header2 = "└──────────────────────────────── ♪ ─┘"
    print(header2.center(50), "\n")
    # Dependiendo al subgénero, se imprime el nombre del artista/grupo y
    # una canción que se encuentre en la base de datos
    if "Dream pop" in listapop:
        for pop1 in sheet.iter_rows(min_row=11, max_row=11, min_col=5, max_col=7, values_only=True):
            listarespop.append(pop1)
            res = " - ".join(pop1)
            print("\t- Dream Pop: ", res)
    if "Power pop" in listapop:
        for pop2 in sheet.iter_rows(min_row=14, max_row=14, min_col=5, max_col=7, values_only=True):
            listarespop.append(pop2)
            res = " - ".join(pop2)
            print("\t- Folk Metal: ", res)
    if "J-pop" in listapop:
        for pop3 in sheet.iter_rows(min_row=17, max_row=17, min_col=5, max_col=7, values_only=True):
            listarespop.append(pop3)
            res = " - ".join(pop3)
            print("\t- J-Pop: ", res)
    print("")
    # Se pregunta si se abre automáticamente el link de una canción a youtube
    print("¿Quieres que te llevemos directo al link de youtube de alguna canción?")
    print("\t A) ¡Sí!")
    print("\t B) No")
    reslink = input()
    while reslink not in val:
        reslink = input("Respuesta desconocida. Sólo introduce A ó B\n")
    if reslink == "A":
        if len(listapop) == 3:
            print("¡Perfecto :D! ¿De cuál subgénero quieres el link?")
            if "Dream pop" and "Power pop" and "J-pop" in listapop:
                print("\tA) Dream Pop")
                print("\tB) Power Pop")
                print("\tC) J-Pop\t")
                reslink1 = input()
                while reslink1 not in val:
                    reslink1 = input("Respuesta desconocida. Sólo introduce A ó B\n")
                # Se abre el link del subgénero seleccionado
                if reslink1 == "A":
                    propop = listarespop[0]
                    webbrowser.open(propop[2])
                    print("¡Esperemos que te guste!:D")
                elif reslink1 == "B":
                    propop = listarespop[1]
                    webbrowser.open(propop[2])
                    print("¡Esperemos que te guste!:D")
                elif reslink1 == "C":
                    propop = listarespop[2]
                    webbrowser.open(propop[2])
                    print("¡Esperemos que te guste!:D")
                print("¿Quieres seguir con el sistema de recomendación o volver al menú?")
                print("\t A) Seguir con el sistema")
                print("\t B) Menú principal")
                popfin = input()
                while popfin not in valink:
                    popfin = input("Respuesta desconocida. Sólo introduce A ó B\n")
                if popfin == "A":
                    preguntasRock()
                if popfin == "B":
                    menu()
        elif len(listapop) == 2:
            print("¡Perfecto :D! ¿De cuál subgénero quieres el link?")
            if "Dream pop" and "Power pop" in listapop:
                print("\tA) Dream pop")
                print("\tB) Power pop")
                reslink1 = input()
                while reslink1 not in valink:
                    reslink1 = input("Respuesta desconocida. Sólo introduce A ó B\n")
                # Se abre el link del subgénero seleccionado
                if reslink1 == "A":
                    propop = listarespop[0]
                    webbrowser.open(propop[2])
                    print("¡Esperemos que te guste!:D")
                elif reslink1 == "B":
                    propop = listarespop[1]
                    webbrowser.open(propop[2])
                    print("¡Esperemos que te guste!:D")
                print("¿Quieres seguir con el sistema de recomendación o volver al menú?")
                print("\t A) Seguir con el sistema")
                print("\t B) Menú principal")
                popfin = input()
                while popfin not in valink:
                    popfin = input("Respuesta desconocida. Sólo introduce A ó B\n")
                if popfin == "A":
                    preguntasRock()
                if popfin == "B":
                    menu()
            elif "Dream pop" and "J-pop" in listapop:
                print("\tA) Dream pop")
                print("\tB) J-pop")
                reslink1 = input()
                while reslink1 not in valink:
                    reslink1 = input("Respuesta desconocida. Sólo introduce A ó B\n")
                # Se abre el link del subgénero seleccionado
                if reslink1 == "A":
                    propop = listarespop[0]
                    webbrowser.open(propop[2])
                    print("¡Esperemos que te guste!:D")
                elif reslink1 == "B":
                    propop = listarespop[1]
                    webbrowser.open(propop[2])
                    print("¡Esperemos que te guste!:D")
                print("¿Quieres seguir con el sistema de recomendación o volver al menú?")
                print("\t A) Seguir con el sistema")
                print("\t B) Menú principal")
                popfin = input()
                while popfin not in valink:
                    popfin = input("Respuesta desconocida. Sólo introduce A ó B\n")
                if popfin == "A":
                    preguntasRock()
                if popfin == "B":
                    menu()
            elif "Power pop" and "J-pop" in listapop:
                print("\tA) Power pop")
                print("\tB) J-pop")
                reslink1 = input()
                while reslink1 not in valink:
                    reslink1 = input("Respuesta desconocida. Sólo introduce A ó B\n")
                # Se abre el link del subgénero seleccionado
                if reslink1 == "A":
                    propop = listarespop[0]
                    webbrowser.open(propop[2])
                    print("¡Esperemos que te guste!:D")
                elif reslink1 == "B":
                    propop = listarespop[1]
                    webbrowser.open(propop[2])
                    print("¡Esperemos que te guste!:D")
                print("¿Quieres seguir con el sistema de recomendación o volver al menú?")
                print("\t A) Seguir con el sistema")
                print("\t B) Menú principal")
                popfin = input()
                while popfin not in valink:
                    popfin = input("Respuesta desconocida. Sólo introduce A ó B\n")
                if popfin == "A":
                    preguntasRock()
                if popfin == "B":
                    menu()
        elif len(listapop) == 1:
            print("¡Perfecto :D! En un momento se abrirá")
            propop = listarespop[0]
            webbrowser.open(propop[2])
            print("¡Esperemos que te guste!:D")

    if reslink == "B":
        print()
    print("¿Quieres seguir con el sistema de recomendación o volver al menú?")
    print("\t A) Seguir con el sistema")
    print("\t B) Menú principal")
    # Dependiendo a la respuesta del usuario, se regresa al menú o se sigue con el sistema
    popfin = input()
    while popfin not in valink:
        popfin = input("Respuesta desconocida. Sólo introduce A ó B\n")
    if popfin == "A":
        preguntasRock()
    if popfin == "B":
        menu()


# Función con las preguntas de Rock
def preguntasRock():
    print("¿Qué tanto te agrada cuándo en las canciones hay predominancia de la guitarra eléctrica,"
          " con canciones de compás 4/4 y una estructura verso-estribillo?\n")
    print("A) ¡Mucho!")
    print("B) ¡Sí! Me gusta")
    print("C) No mucho pero no me desagrada")
    print("D) Prefiero mejor otro tipo de música \n")
    RRespuesta1 = input()
    while RRespuesta1 not in val1:
        RRespuesta1 = input("Respuesta desconocida, introduce sólo A, B, C ó D\n")
    if RRespuesta1 == "A" or RRespuesta1 == "B":
        print("¿Qué opinas de los ritmos ambiciosos, que cuentan con elementos estéticos "
              "y líricos, además de tener influencias del jazz, la música clásica y la música avantgarde?\n")
        print("A = ¡Son de lo mejor que hay!")
        print("B = Creo que son melodías interesantes")
        print("C = Meh, prefiero algo más\n")
        RRespuesta11 = input()
        while RRespuesta11 not in val:
            RRespuesta11 = input("Respuesta desconocida, introduce sólo A, B ó C\n")
        # Se agrega el subgénero a la lista del género
        if RRespuesta11 == "A":
            subgenero7 = "Art rock"
            listarock.append(subgenero7)
        # Se agrega el subgénero a la lista del género
        elif RRespuesta11 == "B":
            subgenero75 = "Art rock"
            listarock.append(subgenero75)

        elif RRespuesta11 == "C":
            print()
        print("¿Qué te parece la música insistentemente alta con letras de protesta contra"
              " la sociedad usando generalmente un lenguaje violento?\n")
        print("A = No puedo vivir sin esa música")
        print("B = Pues, no son de mi estilo, pero me agrada la melodía")
        print("C = Muy loco, mejor otra cosa\n")
        RRespuesta12 = input()
        while RRespuesta12 not in val:
            RRespuesta12 = input("Respuesta desconocida, introduce sólo A, B ó C\n")
        # Se agrega el subgénero a la lista del género
        if RRespuesta12 == "A":
            subgenero8 = "Punk rock"
            listarock.append(subgenero8)
        # Se agrega el subgénero a la lista del género
        elif RRespuesta12 == "B":
            subgenero85 = "Punk rock"
            listarock.append(subgenero85)

        elif RRespuesta12 == "C":
            print()
        print("¿Te gustan las canciones que incorporan instrumentos como la guitarra,"
              " el bajo, la batería, teclados y sintetizadores, además de vocales?\n")
        print("A = ¡No hay nada mejor que estas canciones!")
        print("B = Ocasionalmente las escucho, pero no son de mis favoritas")
        print("C = No me gustan tanto\n")
        RRespuesta13 = input()
        while RRespuesta13 not in val:
            RRespuesta13 = input("Respuesta desconocida, introduce sólo A, B ó C\n")
        # Se agrega el subgénero a la lista del género
        if RRespuesta13 == "A":
            subgenero9 = "Rock independiente"
            listarock.append(subgenero9)
        # Se agrega el subgénero a la lista del género
        elif RRespuesta13 == "B":
            subgenero95 = "Rock independiente"
            listarock.append(subgenero95)

        elif RRespuesta13 == "C":
            print()

        # Se envían a los usuarios a las recomendaciones de canciones y artistas,
        # en el caso que hayan seleccionado A o B en una de las preguntas
        if len(listarock) == 1:
            print("¡Muy bien! Te gustó el siguiente subgénero del rock:")
            print('\t, '.join(listarock))
            print("¡Te recomendaremos el siguiente artista con una canción! Sabemos que te encantará")
            print("")
            resultadosRock()
        elif len(listarock) >= 2:
            print("¡Muy bien! Te recomendamos los siguientes subgéneros del rock:")
            print("", end="\t")
            print(', '.join(listarock))
            print(
                "¡Te recomendamos los siguiente artistas con su determinado subgénero y"
                " una de sus canciones! Sabemos que te encantarán :D")
            print("")
            resultadosRock()

    # Si no hay valores en la lista o se seleccionó "C" o "D" en la preguntas inicial,
        # se pasa a las preguntas del siguiente género
        elif len(listarock) == 0:
            preguntasElectro()
    elif RRespuesta1 == "C" or RRespuesta1 == "D":
        preguntasElectro()


# Función con las recomendaciones resultantes de Rock
def resultadosRock():
    # Se lee la base de datos en la primera hoja
    from openpyxl import load_workbook
    filesheet = "./BaseDeDatos.xlsx"
    wb = load_workbook(filesheet)
    wb.active = 0
    sheet = wb.active
    # Se imprime el título con estilo
    header = "┌─ ♪ ────────────────────────────────┐"
    print("\n", header.center(48), "\n")
    tittle = "¸♬¸ Resultados ¸♬¸"
    print(tittle.center(49), "\n")
    header2 = "└──────────────────────────────── ♪ ─┘"
    print(header2.center(50), "\n")
    # Dependiendo al subgénero, se imprime el nombre del artista/grupo
    # y una canción que se encuentre en la base de datos
    if "Art rock" in listarock:
        for rock1 in sheet.iter_rows(min_row=20, max_row=20, min_col=5, max_col=7, values_only=True):
            listaresrock.append(rock1)
            res = " - ".join(rock1)
            print("\t- Art rock: ", res)
    if "Punk rock" in listarock:
        for rock2 in sheet.iter_rows(min_row=23, max_row=23, min_col=5, max_col=7, values_only=True):
            listaresrock.append(rock2)
            res = " - ".join(rock2)
            print("\t- Punk rock: ", res)
    if "Rock independiente" in listarock:
        for rock3 in sheet.iter_rows(min_row=26, max_row=26, min_col=5, max_col=7, values_only=True):
            listaresrock.append(rock3)
            res = " - ".join(rock3)
            print("\t- Rock independiente: ", res)
    print("")
    # Se pregunta si se abre automáticamente el link de una canción a youtube
    print("¿Quieres que te llevemos directo al link de youtube de alguna canción?")
    print("\t A) ¡Sí!")
    print("\t B) No")
    reslink = input()
    while reslink not in val:
        reslink = input("Respuesta desconocida. Sólo introduce A ó B\n")
    if reslink == "A":
        if len(listarock) == 3:
            print("¡Perfecto :D! ¿De cuál subgénero quieres el link?")
            if "Art rock" and "Punk rock" and "Rock independiente" in listarock:
                print("\tA) Art rock")
                print("\tB) Punk rock")
                print("\tC) Rock independiente")
                reslink1 = input()
                while reslink1 not in val:
                    reslink1 = input("Respuesta desconocida. Sólo introduce A ó B\n")
                # Se abre el link del subgénero seleccionado
                if reslink1 == "A":
                    prorock = listaresrock[0]
                    webbrowser.open(prorock[2])
                    print("¡Esperemos que te guste!:D")
                elif reslink1 == "B":
                    prorock = listaresrock[1]
                    webbrowser.open(prorock[2])
                    print("¡Esperemos que te guste!:D")
                elif reslink1 == "C":
                    prorock = listaresrock[2]
                    webbrowser.open(prorock[2])
                    print("¡Esperemos que te guste!:D")
                print("¿Quieres seguir con el sistema de recomendación o volver al menú?")
                print("\t A) Seguir con el sistema")
                print("\t B) Menú principal")
                rockfin = input()
                while rockfin not in valink:
                    rockfin = input("Respuesta desconocida. Sólo introduce A ó B\n")
                if rockfin == "A":
                    preguntasElectro()
                if rockfin == "B":
                    menu()
        elif len(listarock) == 2:
            print("¡Perfecto :D! ¿De cuál subgénero quieres el link?")
            if "Art rock" and "Punk rock" in listarock:
                print("\tA) Art rock")
                print("\tB) Punk rock")
                reslink1 = input()
                while reslink1 not in valink:
                    reslink1 = input("Respuesta desconocida. Sólo introduce A ó B\n")
                # Se abre el link del subgénero seleccionado
                if reslink1 == "A":
                    prorock = listaresrock[0]
                    webbrowser.open(prorock[2])
                    print("¡Esperemos que te guste!:D")
                elif reslink1 == "B":
                    prorock = listaresrock[1]
                    webbrowser.open(prorock[2])
                    print("¡Esperemos que te guste!:D")
                print("¿Quieres seguir con el sistema de recomendación o volver al menú?")
                print("\t A) Seguir con el sistema")
                print("\t B) Menú principal")
                rockfin = input()
                while rockfin not in valink:
                    rockfin = input("Respuesta desconocida. Sólo introduce A ó B\n")
                if rockfin == "A":
                    preguntasElectro()
                if rockfin == "B":
                    menu()
            elif "Art rock" and "Rock independiente" in listarock:
                print("\tA) Art rock")
                print("\tB) Rock independiente")
                reslink1 = input()
                while reslink1 not in valink:
                    reslink1 = input("Respuesta desconocida. Sólo introduce A ó B\n")
                # Se abre el link del subgénero seleccionado
                if reslink1 == "A":
                    prorock = listaresrock[0]
                    webbrowser.open(prorock[2])
                    print("¡Esperemos que te guste!:D")
                elif reslink1 == "B":
                    prorock = listaresrock[1]
                    webbrowser.open(prorock[2])
                    print("¡Esperemos que te guste!:D")
                print("¿Quieres seguir con el sistema de recomendación o volver al menú?")
                print("\t A) Seguir con el sistema")
                print("\t B) Menú principal\n")
                rockfin = input()
                while rockfin not in valink:
                    rockfin = input("Respuesta desconocida. Sólo introduce A ó B\n")
                if rockfin == "A":
                    preguntasElectro()
                if rockfin == "B":
                    menu()
            elif "Punk rock" and "Rock independiente" in listarock:
                print("\tA) Punk rock")
                print("\tB) Rock independiente")
                reslink1 = input()
                while reslink1 not in valink:
                    reslink1 = input("Respuesta desconocida. Sólo introduce A ó B\n")
                # Se abre el link del subgénero seleccionado
                if reslink1 == "A":
                    prorock = listaresrock[0]
                    webbrowser.open(prorock[2])
                    print("¡Esperemos que te guste!:D")
                elif reslink1 == "B":
                    prorock = listaresrock[1]
                    webbrowser.open(prorock[2])
                    print("¡Esperemos que te guste!:D")
                print("¿Quieres seguir con el sistema de recomendación o volver al menú?")
                print("\t A) Seguir con el sistema")
                print("\t B) Menú principal")
                rockfin = input()
                while rockfin not in valink:
                    rockfin = input("Respuesta desconocida. Sólo introduce A ó B\n")
                if rockfin == "A":
                    preguntasElectro()
                if rockfin == "B":
                    menu()
        elif len(listarock) == 1:
            print("¡Perfecto :D! En un momento se abrirá")
            prorock = listaresrock[0]
            webbrowser.open(prorock[2])
            print("¡Esperemos que te guste!:D")

    if reslink == "B":
        print()
    print("¿Quieres seguir con el sistema de recomendación o volver al menú?")
    print("\t A) Seguir con el sistema")
    print("\t B) Menú principal")
    rockfin = input()
    while rockfin not in valink:
        rockfin = input("Respuesta desconocida. Sólo introduce A ó B\n")
    # Dependiendo a la respuesta del usuario, se regresa al menú o se sigue con el sistema
    if rockfin == "A":
        preguntasElectro()
    if rockfin == "B":
        menu()


# Función con las preguntas de Electrónica
def preguntasElectro():
    print("¿Te gustan las melodías y sonidos, generalmente sin vocales,"
          " que estén procesadas a través de un ordenador?\n")
    print("A) ¡Mucho!")
    print("B) ¡Sí! Me gustan")
    print("C) No mucho pero no me desagradan")
    print("D) Prefiero mejor otro tipo de melodías \n")
    ERespuesta1 = input()
    while ERespuesta1 not in val1:
        ERespuesta1 = input("Respuesta desconocida, introduce sólo A, B, C ó D\n")
    if ERespuesta1 == "A" or ERespuesta1 == "B":
        print("¿Te podría interesar la combinación de la electrónica, el jazz, el hip hop "
              "y principalmente, el swing moderno?\n")
        print("A = Woah, suena súper bien")
        print("B = Suena interesante")
        print("C = Creo que pasaré\n")
        ERespuesta11 = input()
        while ERespuesta11 not in val:
            ERespuesta11 = input("Respuesta desconocida, introduce sólo A, B ó C\n")
        # Se agrega el subgénero a la lista del género
        if ERespuesta11 == "A":
            subgenero10 = "Electro-swing"
            listaelectro.append(subgenero10)
        # Se agrega el subgénero a la lista del género
        elif ERespuesta11 == "B":
            subgenero105 = "Electro-swing"
            listaelectro.append(subgenero105)

        elif ERespuesta11 == "C":
            print()
        print("¿Qué tal suena música que mezcla y fusiona varios subgéneros del rock con"
              " la música electrónica de baile?\n")
        print("A = ¡A darle!")
        print("B = Solo para pasar el rato")
        print("C = Mejor cámbienle a la música\n")
        ERespuesta12 = input()
        while ERespuesta12 not in val:
            ERespuesta12 = input("Respuesta desconocida, introduce sólo A, B ó C\n")
        # Se agrega el subgénero a la lista del género
        if ERespuesta12 == "A":
            subgenero11 = "Dance alternativo"
            listaelectro.append(subgenero11)
        # Se agrega el subgénero a la lista del género
        elif ERespuesta12 == "B":
            subgenero115 = "Dance alternativo"
            listaelectro.append(subgenero115)

        elif ERespuesta12 == "C":
            print()
        print("¿Conoces la música \"House\"?, imagina un sonido más relajante y activo,"
              " con un tempo más bajo ¿te gusta?\n")
        print("A = ¡Siii! Me encantan")
        print("B = Sí la conozco, me gusta algo / Pues, creo que me podría gustar")
        print("C = No la conozco, pero no me interesa / Sí la conozco, pero no me gusta\n")
        ERespuesta13 = input()
        while ERespuesta13 not in val:
            ERespuesta13 = input("Respuesta desconocida, introduce sólo A, B ó C\n")
        # Se agrega el subgénero a la lista del género
        if ERespuesta13 == "A":
            subgenero12 = "Tropical house"
            listaelectro.append(subgenero12)
        # Se agrega el subgénero a la lista del género
        elif ERespuesta13 == "B":
            subgenero125 = "Tropical house"
            listaelectro.append(subgenero125)

        elif ERespuesta13 == "C":
            print()

        # Se envían a los usuarios a las recomendaciones de canciones y artistas,
        # en el caso que hayan seleccionado A o B en una de las preguntas
        if len(listaelectro) == 1:
            print("¡Muy bien! Te gustó el siguiente subgénero de la electónica:")
            print('\t, '.join(listaelectro))
            print("¡Te recomendaremos el siguiente artista con una canción! Sabemos que te encantará")
            print("")
            resultadosElectro()
        elif len(listaelectro) >= 2:
            print("¡Muy bien! Te recomendamos los siguientes subgéneros de la electrónica:")
            print("", end="\t")
            print(', '.join(listaelectro))
            print(
                "¡Te recomendamos los siguiente artistas con su determinado subgénero "
                "y una de sus canciones! Sabemos que te encantarán :D")
            print("")
            resultadosElectro()

    # Si no hay valores en la lista o se seleccionó "C" o "D" en la preguntas inicial,
        # se pasa a las preguntas del siguiente género
        elif len(listaelectro) == 0:
            preguntasHipHop()
    elif ERespuesta1 == "C" or ERespuesta1 == "D":
        preguntasHipHop()


# Función con las recomendaciones resultantes de Electrónica
def resultadosElectro():
    # Se lee la base de datos en la primera hoja
    from openpyxl import load_workbook
    filesheet = "./BaseDeDatos.xlsx"
    wb = load_workbook(filesheet)
    wb.active = 0
    sheet = wb.active
    # Se imprime el título con estilo
    header = "┌─ ♪ ────────────────────────────────┐"
    print("\n", header.center(48), "\n")
    tittle = "¸♬¸ Resultados ¸♬¸"
    print(tittle.center(49), "\n")
    header2 = "└──────────────────────────────── ♪ ─┘"
    print(header2.center(50), "\n")
    # Dependiendo al subgénero, se imprime el nombre del artista/grupo
    # y una canción que se encuentre en la base de datos
    if "Electro-swing" in listaelectro:
        for electro1 in sheet.iter_rows(min_row=29, max_row=29, min_col=5, max_col=7, values_only=True):
            listareselectro.append(electro1)
            res = " - ".join(electro1)
            print("\t- Electro-swing: ", res)
    if "Dance alternativo" in listaelectro:
        for electro2 in sheet.iter_rows(min_row=32, max_row=32, min_col=5, max_col=7, values_only=True):
            listareselectro.append(electro2)
            res = " - ".join(electro2)
            print("\t- Dance alternativo: ", res)
    if "Tropical house" in listaelectro:
        for electro3 in sheet.iter_rows(min_row=35, max_row=35, min_col=5, max_col=7, values_only=True):
            listareselectro.append(electro3)
            res = " - ".join(electro3)
            print("\t- Tropical house: ", res)
    print("")
    # Se pregunta si se abre automáticamente el link de una canción a youtube
    print("¿Quieres que te llevemos directo al link de youtube de alguna canción?")
    print("\t A) ¡Sí!")
    print("\t B) No")
    reslink = input()
    while reslink not in val:
        reslink = input("Respuesta desconocida. Sólo introduce A ó B\n")
    if reslink == "A":
        if len(listaelectro) == 3:
            print("¡Perfecto :D! ¿De cuál subgénero quieres el link?")
            if "Electro-swing" and "Dance alternativo" and "Tropical house" in listaelectro:
                print("\tA) Electro-swing")
                print("\tB) Dance alternativo")
                print("\tC) Tropical house")
                reslink1 = input()
                while reslink1 not in val:
                    reslink1 = input("Respuesta desconocida. Sólo introduce A ó B\n")
                # Se abre el link del subgénero seleccionado
                if reslink1 == "A":
                    proelectro = listareselectro[0]
                    webbrowser.open(proelectro[2])
                    print("¡Esperemos que te guste!:D")
                elif reslink1 == "B":
                    proelectro = listareselectro[1]
                    webbrowser.open(proelectro[2])
                    print("¡Esperemos que te guste!:D")
                elif reslink1 == "C":
                    proelectro = listareselectro[2]
                    webbrowser.open(proelectro[2])
                    print("¡Esperemos que te guste!:D")
                print("¿Quieres seguir con el sistema de recomendación o volver al menú?")
                print("\t A) Seguir con el sistema")
                print("\t B) Menú principal")
                electrofin = input()
                while electrofin not in valink:
                    electrofin = input("Respuesta desconocida. Sólo introduce A ó B\n")
                if electrofin == "A":
                    preguntasHipHop()
                if electrofin == "B":
                    menu()
        elif len(listaelectro) == 2:
            print("¡Perfecto :D! ¿De cuál subgénero quieres el link?")
            if "Electro-swing" and "Dance alternativo" in listaelectro:
                print("\tA) Electro-swing")
                print("\tB) Dance alternativo")
                reslink1 = input()
                while reslink1 not in valink:
                    reslink1 = input("Respuesta desconocida. Sólo introduce A ó B\n")
                # Se abre el link del subgénero seleccionado
                if reslink1 == "A":
                    proelectro = listareselectro[0]
                    webbrowser.open(proelectro[2])
                    print("¡Esperemos que te guste!:D")
                elif reslink1 == "B":
                    proelectro = listareselectro[1]
                    webbrowser.open(proelectro[2])
                    print("¡Esperemos que te guste!:D")
                print("¿Quieres seguir con el sistema de recomendación o volver al menú?")
                print("\t A) Seguir con el sistema")
                print("\t B) Menú principal")
                electrofin = input()
                while electrofin not in valink:
                    electrofin = input("Respuesta desconocida. Sólo introduce A ó B\n")
                if electrofin == "A":
                    preguntasHipHop()
                if electrofin == "B":
                    menu()
            elif "Electro-swing" and "Tropical house" in listaelectro:
                print("\tA) Electro-swing")
                print("\tB) Tropical house")
                reslink1 = input()
                while reslink1 not in valink:
                    reslink1 = input("Respuesta desconocida. Sólo introduce A ó B\n")
                # Se abre el link del subgénero seleccionado
                if reslink1 == "A":
                    proelectro = listareselectro[0]
                    webbrowser.open(proelectro[2])
                    print("¡Esperemos que te guste!:D")
                elif reslink1 == "B":
                    proelectro = listareselectro[1]
                    webbrowser.open(proelectro[2])
                    print("¡Esperemos que te guste!:D")
                print("¿Quieres seguir con el sistema de recomendación o volver al menú?")
                print("\t A) Seguir con el sistema")
                print("\t B) Menú principal\n")
                electrofin = input()
                while electrofin not in valink:
                    electrofin = input("Respuesta desconocida. Sólo introduce A ó B\n")
                if electrofin == "A":
                    preguntasHipHop()
                if electrofin == "B":
                    menu()
            elif "Dance alternativo" and "Tropical house" in listaelectro:
                print("\tA) Dance alternativo")
                print("\tB) Tropical house")
                reslink1 = input()
                while reslink1 not in valink:
                    reslink1 = input("Respuesta desconocida. Sólo introduce A ó B\n")
                # Se abre el link del subgénero seleccionado
                if reslink1 == "A":
                    proelectro = listareselectro[0]
                    webbrowser.open(proelectro[2])
                    print("¡Esperemos que te guste!:D")
                elif reslink1 == "B":
                    prorock = listareselectro[1]
                    webbrowser.open(prorock[2])
                    print("¡Esperemos que te guste!:D")
                print("¿Quieres seguir con el sistema de recomendación o volver al menú?")
                print("\t A) Seguir con el sistema")
                print("\t B) Menú principal")
                electrofin = input()
                while electrofin not in valink:
                    electrofin = input("Respuesta desconocida. Sólo introduce A ó B\n")
                if electrofin == "A":
                    preguntasHipHop()
                if electrofin == "B":
                    menu()
        elif len(listaelectro) == 1:
            print("¡Perfecto :D! En un momento se abrirá")
            proelectro = listareselectro[0]
            webbrowser.open(proelectro[2])
            print("¡Esperemos que te guste!:D")

    if reslink == "B":
        print()
    print("¿Quieres seguir con el sistema de recomendación o volver al menú?")
    print("\t A) Seguir con el sistema")
    print("\t B) Menú principal")
    electrofin = input()
    while electrofin not in valink:
        electrofin = input("Respuesta desconocida. Sólo introduce A ó B\n")
    # Dependiendo a la respuesta del usuario, se regresa al menú o se sigue con el sistema
    if electrofin == "A":
        preguntasHipHop()
    if electrofin == "B":
        menu()


# Función con las preguntas de Hip-Hop
def preguntasHipHop():
    print("¿Disfrutas cuando hay secciones en una canción en las que el cantante incorpora rimas "
          "y una fluidez rítmica a los versos?\n")
    print("A) ¡Me encanta!")
    print("B) ¡Sí! Me gusta")
    print("C) Me agrada pero no es ")
    print("D) Prefiero mejor otro tipo de melodías \n")
    HRespuesta1 = input()
    while HRespuesta1 not in val1:
        HRespuesta1 = input("Respuesta desconocida, introduce sólo A, B, C ó D\n")
    if HRespuesta1 == "A" or HRespuesta1 == "B":
        print("¿Qué opinas de la música \"chillout\" fusionada con el hip hop,"
              " en donde se emplea una baja fidelidad durante la grabación?\n")
        print("A = Me encanta escuchala frecuentemente")
        print("B = A veces la escucho, pero no muy seguido")
        print("C = Nah, no son de mi estilo\n")
        HRespuesta11 = input()
        while HRespuesta11 not in val:
            HRespuesta11 = input("Respuesta desconocida, introduce sólo A, B ó C\n")
        # Se agrega el subgénero a la lista del género
        if HRespuesta11 == "A":
            subgenero13 = "LoFi"
            listahip.append(subgenero13)
        # Se agrega el subgénero a la lista del género
        elif HRespuesta11 == "B":
            subgenero135 = "LoFi"
            listahip.append(subgenero135)

        elif HRespuesta11 == "C":
            print()
        print("¿Te llama la atención la música que cubre varios estilos de rap hechos en Latinoamérica y España?\n")
        print("A = ¡Bastante!")
        print("B = Podría interesarme")
        print("C = No mucho, no\n")
        HRespuesta12 = input()
        while HRespuesta12 not in val:
            HRespuesta12 = input("Respuesta desconocida, introduce sólo A, B ó C\n")
        # Se agrega el subgénero a la lista del género
        if HRespuesta12 == "A":
            subgenero14 = "Latino"
            listahip.append(subgenero14)
        # Se agrega el subgénero a la lista del género
        elif HRespuesta12 == "B":
            subgenero145 = "Latino"
            listahip.append(subgenero145)

        elif HRespuesta12 == "C":
            print()
        print("¿Qué te parece la música que incorpora melodías lentas e \"hipnotizantes\", un bajo grave, v"
              "oces femeninas de fondo y el uso de sintetizadores?\n")
        print("A = Woah, se escucha padre")
        print("B = Podría gustarme, tal vez")
        print("C = No creo que sea de mi tipo\n")
        HRespuesta13 = input()
        while HRespuesta13 not in val:
            HRespuesta13 = input("Respuesta desconocida, introduce sólo A, B ó C\n")
        # Se agrega el subgénero a la lista del género
        if HRespuesta13 == "A":
            subgenero15 = "gfunk"
            listahip.append(subgenero15)
        # Se agrega el subgénero a la lista del género
        elif HRespuesta13 == "B":
            subgenero155 = "gfunk"
            listahip.append(subgenero155)

        elif HRespuesta13 == "C":
            print()

        # Se envían a los usuarios a las recomendaciones de canciones y artistas,
        # en el caso que hayan seleccionado A o B en una de las preguntas
        if len(listahip) == 1:
            print("¡Muy bien! Te gustó el siguiente subgénero del hip-hop:")
            print('\t, '.join(listahip))
            print("¡Te recomendaremos el siguiente artista con una canción! Sabemos que te encantará")
            print("")
            resultadosHipHop()
        elif len(listahip) >= 2:
            print("¡Muy bien! Te recomendamos los siguientes subgéneros del hip-hop:")
            print("", end="\t")
            print(', '.join(listahip))
            print("¡Te recomendamos los siguiente artistas con su determinado subgénero y una de sus canciones!"
                  " Sabemos que te encantarán :D")
            print("")
            resultadosHipHop()

    # Si no hay valores en la lista o se seleccionó "C" o "D" en la preguntas inicial, se pasa directamente al menú
        elif len(listahip) == 0:
            print("Fin del sistema de recomendación")
            print("¡Espero te hayan gustado las recomendaciones que te dimos!")
            print("Te enviaremos al menú principal")
            menu()



    elif HRespuesta1 == "C" or HRespuesta1 == "D":
        print("Fin del sistema de recomendación")
        print("¡Espero te hayan gustado las recomendaciones que te dimos!")
        print("Te enviaremos al menú principal")
        menu()


# Función con las recomendaciones resultantes de Hip-Hop
def resultadosHipHop():
    # Se lee la base de datos en la primera hoja
    from openpyxl import load_workbook
    filesheet = "./BaseDeDatos.xlsx"
    wb = load_workbook(filesheet)
    wb.active = 0
    sheet = wb.active
    # Se imprime el título con estilo
    header = "┌─ ♪ ────────────────────────────────┐"
    print("\n", header.center(48), "\n")
    tittle = "¸♬¸ Resultados ¸♬¸"
    print(tittle.center(49), "\n")
    header2 = "└──────────────────────────────── ♪ ─┘"
    print(header2.center(50), "\n")
    # Dependiendo al subgénero, se imprime el nombre del artista/grupo
    # y una canción que se encuentre en la base de datos
    if "LoFi" in listahip:
        for hip1 in sheet.iter_rows(min_row=38, max_row=38, min_col=5, max_col=7, values_only=True):
            listareship.append(hip1)
            res = " - ".join(hip1)
            print("\t- LoFi: ", res)
    if "Latino" in listahip:
        for hip2 in sheet.iter_rows(min_row=41, max_row=41, min_col=5, max_col=7, values_only=True):
            listareship.append(hip2)
            res = " - ".join(hip2)
            print("\t- Latino: ", res)
    if "gfunk" in listahip:
        for hip3 in sheet.iter_rows(min_row=44, max_row=44, min_col=5, max_col=7, values_only=True):
            listareship.append(hip3)
            res = " - ".join(hip3)
            print("\t- G-Funk: ", res)
    print("")
    # Se pregunta si se abre automáticamente el link de una canción a youtube
    print("¿Quieres que te llevemos directo al link de youtube de alguna canción?")
    print("\t A) ¡Sí!")
    print("\t B) No")
    reslink = input()
    while reslink not in val:
        reslink = input("Respuesta desconocida. Sólo introduce A ó B\n")
    if reslink == "A":
        if len(listahip) == 3:
            print("¡Perfecto :D! ¿De cuál subgénero quieres el link?")
            if "LoFi" and "Latino" and "gfunk" in listahip:
                print("\tA) LoFi")
                print("\tB) Latino")
                print("\tC) G-Funk")
                reslink1 = input()
                while reslink1 not in val:
                    reslink1 = input("Respuesta desconocida. Sólo introduce A ó B\n")
                # Se abre el link del subgénero seleccionado
                if reslink1 == "A":
                    prohip = listareship[0]
                    webbrowser.open(prohip[2])
                    print("¡Esperemos que te guste!:D")
                # Se abre el link del subgénero seleccionado
                elif reslink1 == "B":
                    prohip = listareship[1]
                    webbrowser.open(prohip[2])
                    print("¡Esperemos que te guste!:D")
                # Se abre el link del subgénero seleccionado
                elif reslink1 == "C":
                    prohip = listareship[2]
                    webbrowser.open(prohip[2])
                    print("¡Esperemos que te guste!:D")
                print("Fin del sistema de recomendación")
                print("¡Espero te hayan gustado las recomendaciones que te dimos!")
                print("Te enviaremos al menú principal")
                menu()
        elif len(listahip) == 2:
            print("¡Perfecto :D! ¿De cuál subgénero quieres el link?")
            if "LoFi" and "Latino" in listahip:
                print("¿De cuál subgénero quieres el link?")
                print("\tA) LoFi")
                print("\tB) Latino")
                reslink1 = input()
                while reslink1 not in valink:
                    reslink1 = input("Respuesta desconocida. Sólo introduce A ó B\n")
                # Se abre el link del subgénero seleccionado
                if reslink1 == "A":
                    prohip = listareship[0]
                    webbrowser.open(prohip[2])
                    print("¡Esperemos que te guste!:D")
                # Se abre el link del subgénero seleccionado
                elif reslink1 == "B":
                    prohip = listareship[1]
                    webbrowser.open(prohip[2])
                    print("¡Esperemos que te guste!:D")
                print("Fin del sistema de recomendación")
                print("¡Espero te hayan gustado las recomendaciones que te dimos!")
                print("Te enviaremos al menú principal")
                menu()
            elif "LoFi" and "gfunk" in listahip:
                print("\tA) LoFi")
                print("\tB) G-Funk")
                reslink1 = input()
                while reslink1 not in valink:
                    reslink1 = input("Respuesta desconocida. Sólo introduce A ó B\n")
                # Se abre el link del subgénero seleccionado
                if reslink1 == "A":
                    prohip = listareship[0]
                    webbrowser.open(prohip[2])
                    print("¡Esperemos que te guste!:D")
                elif reslink1 == "B":
                    prohip = listareship[1]
                    webbrowser.open(prohip[2])
                    print("¡Esperemos que te guste!:D")
                print("Fin del sistema de recomendación")
                print("¡Espero te hayan gustado las recomendaciones que te dimos!")
                print("Te enviaremos al menú principal")
                menu()
            elif "Latino" and "gfunk" in listahip:
                print("\tA) Latino")
                print("\tB) G-Funk")
                reslink1 = input()
                while reslink1 not in valink:
                    reslink1 = input("Respuesta desconocida. Sólo introduce A ó B\n")
                # Se abre el link del subgénero seleccionado
                if reslink1 == "A":
                    prohip = listareship[0]
                    webbrowser.open(prohip[2])
                    print("¡Esperemos que te guste!:D")
                elif reslink1 == "B":
                    prohip = listareship[1]
                    webbrowser.open(prohip[2])
                    print("¡Esperemos que te guste!:D")
                print("Fin del sistema de recomendación")
                print("¡Espero te hayan gustado las recomendaciones que te dimos!")
                print("Te enviaremos al menú principal")
                menu()
        elif len(listahip) == 1:
            print("¡Perfecto :D! En un momento se abrirá")
            prohip = listareship[0]
            webbrowser.open(prohip[2])
            print("¡Esperemos que te guste!:D")
    # Al finalizar el sistema de recomendación, se le envía directamente al menú principal.
        print("Fin del sistema de recomendación")
        print("¡Espero te hayan gustado las recomendaciones que te dimos!")
        print("Te enviaremos al menú principal")
        menu()

    if reslink == "B":
        print()
    print("Fin del sistema de recomendación")
    print("¡Espero te hayan gustado las recomendaciones que te dimos!")
    print("Te enviaremos al menú principal")
    menu()


# Ver Recomendaciones anteriores (se leen del archivo)
def recoant():
    # Se le da formato al título
    header = "┌─ ♪ ────────────────────────────────┐"
    print("\n", header.center(50), "\n")
    titulo = "¸♬¸ Recomendaciones anteriores ¸♬¸"
    print(titulo.center(50), "\n")
    header2 = "└──────────────────────────────── ♪ ─┘"
    print(header2.center(50), "\n")
    # Se pide el nombre de usuario para poder leer sus datos
    print("Ingresa tu nombre de usuario de nuevo, por favor :D")
    nomu = input("Nombre de usuario: ")
    # Se valida que el nombre se escriba correctamente
    while nomu not in leernombreusuario():
        print("\nAsegúrate de escribir correctamente tu usuario. Inténtalo de nuevo")
        nomu = input("\tNombre de usuario: ")
    # Se busca el índice en que se encuentra dentro de la lista de usuarios para ese poder leer esas recomendaciones
    renglonusuario = leernombreusuario().index(nomu)
    # Se abre el archivo
    from openpyxl import load_workbook
    filesheet = "./BaseDeDatos.xlsx"
    wb = load_workbook(filesheet)
    # Se lee la hoja 2, "Usuarios" (índice 1)
    wb.active = 1
    sheet = wb.active
    # Se lee el renglón de la hoja 2 en donde este el nombre de usuario y se hace una lista con todos sus valores
    rengloncompleto = []
    row = sheet[renglonusuario+1]
    for r in row:
        rengloncompleto.append(r.value)
    wb.close()
    # Si hay un valor vacío en la lista, se pide hacer el test o regresar al menú
    if None in rengloncompleto:
        print("\n\tDebes hacer el test antes de poder acceder a esta función.")
        Resp1 = ""
        while not ("A" <= Resp1 <= "B"):
            print("Presiona 'A' para proceder al test o 'B' para regresar al Menú", end="")
            Resp1 = input(": ")
            if Resp1 == "A":
                print("Se procede al test")
                # Insertar función de 'Preguntas'
                # !!! Mientras tanto, se termina el programa
                exit()
            else:
                # Se procede al menú
                menu()
                break
    # Se obtiene una sublista a partir del índice 3, en dónde empiezan las recomendaciones pasadas
    recomendacionesp_lista = rengloncompleto[3:]
    # Se imprime con formato
    tablarecop(recomendacionesp_lista[0], recomendacionesp_lista[1], recomendacionesp_lista[2],
               recomendacionesp_lista[3], recomendacionesp_lista[4])
    Resp1 = ""
    while not ("A" <= Resp1 <= "B"):
        print("\nPresiona 'A' para proceder al test o 'B' para regresar al Menú", end="")
        Resp1 = input(": ")
        if Resp1 == "A":
            print("Se procede al test")
            preguntas()
        else:
            # Se procede al menú
            menu()


# Para dar formato estético a la tabla de recomendaciones pasadas
def tablarecop(P1, P2, P3, P4, P5):
    print()
    print("♬ GÉNERO ♬".center(20, " "), end="")
    print(" | ", end="")
    print("♬ SUBGÉNERO ♬".center(25, " "), end="")
    print(" | ", end="")
    print("♬ ARTISTA/GRUPO ♬".center(40, " "), end="")
    print(" | ", end="")
    print("♬ CANCIÓN ♬".center(30, " "), end="")
    print(" | ", end="")
    print("♬ LINK ♬".center(25, " "))
    print(P1.center(21, " "), end="")
    print(" | ", end="")
    print(P2.center(26, " "), end="")
    print(" | ", end="")
    print(P3.center(41, " "), end="")
    print(" | ", end="")
    print(P4.center(31, " "), end="")
    print(" | ", end="")
    print(P5.center(26, " "))


def ingresarmusicnueva1():
    # Se da formato al título
    header = "┌─ ♪ ────────────────────────────────┐"
    print("\n", header.center(50), "\n")
    titulo = "¸♬¸ Ingresar música nueva ¸♬¸"
    print(titulo.center(50), "\n")
    header2 = "└──────────────────────────────── ♪ ─┘"
    print(header2.center(50), "\n")
    # Se piden los datos, se valida cada uno con la función nuevodato
    print("Ingresa el género nuevo o 'X' para regresar al Menú", end="")
    newgen = input(": ")
    nuevodato(newgen)
    print("Ingresa el subgénero nuevo o 'X' para regresar al Menú", end="")
    newsub = input(": ")
    nuevodato(newsub)
    print("Ingresa el artista nuevo o 'X' para regresar al Menú", end="")
    newart = input(": ")
    nuevodato(newart)
    print("Ingresa el grupo nuevo o 'X' para regresar al Menú", end="")
    newgrup = input(": ")
    nuevodato(newgrup)
    print("Ingresa el nombre de la nueva canción o 'X' para regresar al Menú", end="")
    newcanci = input(": ")
    nuevodato(newcanci)
    print("Ingresa el enlace de YouTube a la nueva canción o 'X' para regresar al Menú", end="")
    newlink = input(": ")
    if newlink == "X":
        menu()
    while len(newlink) < 8:
        print("El link debe ser más largo.")
        newlink = input("\tInténtalo de nuevo: ")
    # Se abre la base de datos
    from openpyxl import load_workbook
    filesheet = "./BaseDeDatos.xlsx"
    wb = load_workbook(filesheet)
    # Se lee la hoja 3, "Música Usuarios" (índice 2)
    wb.active = 2
    sheet = wb.active
    # Se busca cuál el último renglón para escribir en él
    siguienteinsert = sheet.max_row
    # Se crea una lista para ingresar los datos
    musicanueva = []
    musicanueva.append(siguienteinsert)
    musicanueva.append(newgen)
    musicanueva.append(newsub)
    musicanueva.append(newart)
    musicanueva.append(newgrup)
    musicanueva.append(newcanci)
    musicanueva.append(newlink)
    # Se graban los datos en el archivo
    try:
        sheet.append(musicanueva)
        wb.save("./BaseDeDatos.xlsx")
        print("\nDatos registrados exitosamente :D")
    except:
        print("\nOcurrió un error al escribir :/")
    finally:
        # Se cierra el archivo
        wb.close()
        # Se muestran los datos agregados
        print("\nTu contribución queda así:")
        print("\t♬ Nuevo género:", newgen)
        print("\t♬ Nuevo subgénero:", newsub)
        print("\t♬ Nuevo artista:", newart)
        print("\t♬ Nuevo grupo:", newgrup)
        print("\t♬ Nueva canción:", newcanci)
        print("\t♬ Enlace a canción:", newlink)
        # Se pregunta si se quiere modificar algún dato y se valida que la respuesta sea A o B
        print("\n¿Quieres cambiar algo? A = Si | B = No", end="")
        cambio = input(": ")
        while not ("A" <= cambio <= "B"):
            print("Respuesta desconocida, sólo introduce 'A' o 'B'")
        # Si la respuesta es B se procede al menú
        if cambio == "B":
            menu()
        # Si la respuesta es A se despliega las opciones para modificar
        else:
            print("¿Que quieres cambiar?")
            print("A = Género")
            print("B = Subgénero")
            print("C = Artista")
            print("D = Grupo")
            print("E = Canción")
            print("F = Enlace")
            # Se pide que se seleccione una opción y se valida que la respuesta sea una letra de la A a la F
            cambiar = input("\n\tCambiar opción: ")
            while not ("A" <= cambiar <= "F"):
                print("Respuesta desconocida. Introduce una opción válida", end="")
                cambiar = input(": ")
            # Dependiendo de la opción elegida se busca su ubicación en la lista que corresponde
            try:
                if cambiar == "A":
                    print("Ingresa nuevo género o 'X' para cancelar", end="")
                    cambionewgen = input(": ")
                    cambiardatos(2, cambionewgen)
                if cambiar == "B":
                    print("Ingresa nuevo subgénero o 'X' para cancelar", end="")
                    cambionewsub = input(": ")
                    cambiardatos(3, cambionewsub)
                if cambiar == "C":
                    print("Ingresa nuevo artista o 'X' para cancelar", end="")
                    cambionewart = input(": ")
                    cambiardatos(4, cambionewart)
                if cambiar == "D":
                    print("Ingresa nuevo grupo o 'X' para cancelar", end="")
                    cambionewgrup = input(": ")
                    cambiardatos(5, cambionewgrup)
                if cambiar == "E":
                    print("Ingresa nuevo nombre de canción o 'X' para cancelar", end="")
                    cambionewcanci = input(": ")
                    cambiardatos(6, cambionewcanci)
                if cambiar == "F":
                    print("Ingresa nuevo género o 'X' para cancelar", end="")
                    cambionewlink = input(": ")
                    cambiardatos(7, cambionewlink)
            except:
                print("\nAlgo salió mal D:")
            finally:
                wb.close()
                print("\nSe procede al menú")
                menu()


# Validar que si el dato es 'X' se regrese al menú o que el dato ingresado tenga más de dos caracteres
def nuevodato(dato):
    if dato == "X":
        menu()
    while len(dato) < 2:
        print("Los nuevos datos deben contener mínimo 2 caracteres.")
        dato = input("\tInténtalo de nuevo: ")


# Leer géneros ingresados por usuarios del archivo
def leergenerosusu():
    from openpyxl import load_workbook
    filesheet = "./BaseDeDatos.xlsx"
    wb = load_workbook(filesheet)
    # Se lee la hoja 3, "Música Usuarios" (índice 2)
    wb.active = 2
    sheet = wb.active
    # Lee columna B (géneros) de la hoja 3 y hace una lista con esos valores
    col = sheet["B"]
    generosusu_lista = []
    for c in col:
        generosusu_lista.append(c.value)
    wb.close()
    return generosusu_lista


# Leer subgéneros ingresados por usuarios del archivo
def leersubgenerosusu():
    from openpyxl import load_workbook
    filesheet = "./BaseDeDatos.xlsx"
    wb = load_workbook(filesheet)
    # Se lee la hoja 3, "Música Usuarios" (índice 2)
    wb.active = 2
    sheet = wb.active
    # Lee columna C (subgéneros) de la hoja 3 y hace una lista con esos valores
    col = sheet["C"]
    subgenerosusu_lista = []
    for c in col:
        subgenerosusu_lista.append(c.value)
    wb.close()
    return subgenerosusu_lista


# Leer artistas ingresados por usuarios del archivo
def leerartistasusu():
    from openpyxl import load_workbook
    filesheet = "./BaseDeDatos.xlsx"
    wb = load_workbook(filesheet)
    # Se lee la hoja 3, "Música Usuarios" (índice 2)
    wb.active = 2
    sheet = wb.active
    # Lee columna D (artistas) de la hoja 3 y hace una lista con esos valores
    col = sheet["D"]
    artistasusu_lista = []
    for c in col:
        artistasusu_lista.append(c.value)
    wb.close()
    return artistasusu_lista


# Leer grupos ingresados por usuarios del archivo
def leergruposusu():
    from openpyxl import load_workbook
    filesheet = "./BaseDeDatos.xlsx"
    wb = load_workbook(filesheet)
    # Se lee la hoja 3, "Música Usuarios" (índice 2)
    wb.active = 2
    sheet = wb.active
    # Lee columna E (grupos) de la hoja 3 y hace una lista con esos valores
    col = sheet["E"]
    gruposusu_lista = []
    for c in col:
        gruposusu_lista.append(c.value)
    wb.close()
    return gruposusu_lista


# Leer canciones ingresados por usuarios del archivo
def leercancionessusu():
    from openpyxl import load_workbook
    filesheet = "./BaseDeDatos.xlsx"
    wb = load_workbook(filesheet)
    # Se lee la hoja 3, "Música Usuarios" (índice 2)
    wb.active = 2
    sheet = wb.active
    # Lee columna F (canciones) de la hoja 3 y hace una lista con esos valores
    col = sheet["F"]
    canionesusu_lista = []
    for c in col:
        canionesusu_lista.append(c.value)
    wb.close()
    return canionesusu_lista


# Leer enlaces ingresados por usuarios del archivo
def leerenlacesusu():
    from openpyxl import load_workbook
    filesheet = "./BaseDeDatos.xlsx"
    wb = load_workbook(filesheet)
    # Se lee la hoja 3, "Música Usuarios" (índice 2)
    wb.active = 2
    sheet = wb.active
    # Lee columna G (links) de la hoja 3 y hace una lista con esos valores
    col = sheet["G"]
    enlacesusu_lista = []
    for c in col:
        enlacesusu_lista.append(c.value)
    wb.close()
    return enlacesusu_lista


def cambiardatos(col, newdato):
    from openpyxl import load_workbook
    filesheet = "./BaseDeDatos.xlsx"
    wb = load_workbook(filesheet)
    # Se lee la hoja 3, "Música Usuarios" (índice 2)
    wb.active = 2
    sheet = wb.active
    # Si el dato ingresado es X, se regresa al menu
    if newdato == "X":
        menu()
    # Sino, se verifica la longitud del dato con la funcion nuevo dato
    else:
        nuevodato(newdato)
    # Se busca en la hoja el dato que se quiere borrar y se sustituye
    renglon = sheet.max_row
    sheet.cell(row=renglon, column=col).value = newdato
    # Se guardan los cambios y se cierra el archivo
    try:
        wb.save("./BaseDeDatos.xlsx")
        print("\nDatos registrados exitosamente :D")
    except:
        print("\nOcurrió un error al escribir :/")
    finally:
        wb.close()


def ingresarmusicnueva2():
    # Se da formato al título
    header = "┌─ ♪ ────────────────────────────────┐"
    print("\n", header.center(50), "\n")
    titulo = "¸♬¸ Ingresar música ¸♬¸"
    print(titulo.center(50), "\n")
    header2 = "└──────────────────────────────── ♪ ─┘"
    print(header2.center(50), "\n")
    # Se pregunta a que género se quiere agregar, se busca en que renglón se va a grabar con la función renglonaescribir
    print("¿A qué género quieres agregar?", end="")
    gen = input(": ")
    reng = renglonaescribir(gen)
    # Se piden los datos, se valida cada uno con la función nuevodato
    print("Ingresa el subgénero nuevo o 'X' para regresar al Menú", end="")
    subgennew = input(": ")
    nuevodato(subgennew)
    print("Ingresa el artista nuevo o 'X' para regresar al Menú", end="")
    artnew = input(": ")
    nuevodato(artnew)
    print("Ingresa el grupo nuevo o 'X' para regresar al Menú", end="")
    grupnew = input(": ")
    nuevodato(grupnew)
    print("Ingresa el nombre de la nueva canción o 'X' para regresar al Menú", end="")
    cancinew = input(": ")
    nuevodato(cancinew)
    print("Ingresa el enlace de YouTube a la nueva canción o 'X' para regresar al Menú", end="")
    linknew = input(": ")
    if linknew == "X":
        menu()
    while len(linknew) < 8:
        print("El link debe ser más largo.")
        linknew = input("\tInténtalo de nuevo: ")
    # Se abre la base de datos
    from openpyxl import load_workbook
    filesheet = "./BaseDeDatos.xlsx"
    wb = load_workbook(filesheet)
    # Se lee la hoja 3, "Música Usuarios" (índice 2)
    wb.active = 2
    sheet = wb.active
    # Se graban los datos de acuerdo a sus caracterísitcas
    try:
        sheet.cell(row=reng, column=3).value = subgennew
        sheet.cell(row=reng, column=4).value = artnew
        sheet.cell(row=reng, column=5).value = grupnew
        sheet.cell(row=reng, column=6).value = cancinew
        sheet.cell(row=reng, column=7).value = linknew
        wb.save("./BaseDeDatos.xlsx")
        print("\nDatos registrados exitosamente :D")
    except:
        print("\nOcurrió un error al escribir :/")
    finally:
        # Se cierra el archivo
        wb.close()
        print("Se procede al menú")
        menu()


# Busca en que renglón se va a grabar dependiendo del género al que se quiere contribuir
def renglonaescribir(genero):
    renglon = int
    if genero == "Metal" or "metal":
        renglon = 2
    elif genero == "Pop" or "pop":
        renglon = 3
    elif genero == "Rock" or "rock":
        renglon = 4
    elif genero == "Electrónica" or "electrónica":
        renglon = 5
    elif genero == "Hip-Hop" or "hip-hop":
        renglon = 6
    return renglon


# Programa: Tunes For You
# El programa se ejecuta con base en las funciones
if saludo() == "A":
    inicio_sesion()
else:
    print("\nDebes crear una cuenta")
    crearcuenta()
